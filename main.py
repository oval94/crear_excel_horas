import os
import sys
import json
import re
import shutil
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

APP_TITLE = "Crear Hoja de Horas"
DEFAULT_BASE_NAME = "Hores i Kilometres"
DEFAULT_RANGE = "B9:H18"

def get_app_dir():
    # Soporta ejecución normal y empaquetada con PyInstaller
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()
CONFIG_PATH = os.path.join(APP_DIR, "config.json")

def load_config():
    cfg = {
        "ruta_trabajo": "",
        "nombre_base": DEFAULT_BASE_NAME,
        "rango_borrar": DEFAULT_RANGE,
        "nombre_hoja": ""
    }
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                cfg.update({k: v for k, v in data.items() if k in cfg})
        except Exception:
            pass
    return cfg

def save_config(cfg):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showwarning(APP_TITLE, f"No se pudo guardar la configuración:\n{e}")

def yyyymm_to_next(yyyymm: str) -> str:
    year = int(yyyymm[:4])
    month = int(yyyymm[4:6])
    if month == 12:
        return f"{year + 1:04d}01"
    return f"{year:04d}{month + 1:02d}"

def find_latest_file(folder: str, base_name: str):
    pattern = re.compile(rf"^(\d{{6}})_{re.escape(base_name)}\.xlsx$", re.IGNORECASE)
    latest_yyyymm = None
    latest_path = None

    try:
        for fname in os.listdir(folder):
            m = pattern.match(fname)
            if m:
                yyyymm = m.group(1)
                y = int(yyyymm[:4]); mth = int(yyyymm[4:6])
                if 1 <= mth <= 12:
                    if (latest_yyyymm is None) or (yyyymm > latest_yyyymm):
                        latest_yyyymm = yyyymm
                        latest_path = os.path.join(folder, fname)
    except FileNotFoundError:
        return (None, None)

    return (latest_path, latest_yyyymm) if latest_path else (None, None)

def clear_range_in_excel(xlsx_path: str, range_a1: str, sheet_name: str | None):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    min_col, min_row, max_col, max_row = range_boundaries(range_a1)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None

    wb.save(xlsx_path)
    wb.close()

def create_next_file(cfg):
    folder = cfg["ruta_trabajo"].strip()
    base = cfg["nombre_base"].strip() or DEFAULT_BASE_NAME
    rng = cfg["rango_borrar"].strip() or DEFAULT_RANGE
    sheet = cfg["nombre_hoja"].strip() or None

    if not folder or not os.path.isdir(folder):
        raise ValueError("Debes indicar una carpeta válida donde están los Excel.")

    latest_path, latest_yyyymm = find_latest_file(folder, base)
    if not latest_path:
        raise FileNotFoundError(
            f"No se encontró ningún archivo con el patrón YYYYMM_{base}.xlsx en:\n{folder}\n"
            f"Crea uno inicial (p. ej. 202602_{base}.xlsx) y vuelve a intentarlo."
        )

    next_yyyymm = yyyymm_to_next(latest_yyyymm)
    dst_name = f"{next_yyyymm}_{base}.xlsx"
    dst_path = os.path.join(folder, dst_name)

    if os.path.exists(dst_path):
        raise FileExistsError(f"Ya existe el archivo destino:\n{dst_path}")

    shutil.copyfile(latest_path, dst_path)
    clear_range_in_excel(dst_path, rng, sheet)

    return dst_path


# --------------------------------------------------------
#                INTERFAZ TKINTER
# --------------------------------------------------------

class App(tk.Tk):
    def __init__(self, cfg):
        super().__init__()
        self.title(APP_TITLE)
        self.resizable(False, False)

        self.cfg = cfg
        pad = {"padx": 8, "pady": 6}

        # Ruta de trabajo
        tk.Label(self, text="Carpeta de trabajo:").grid(row=0, column=0, sticky="w", **pad)
        self.var_ruta = tk.StringVar(value=cfg["ruta_trabajo"])
        tk.Entry(self, textvariable=self.var_ruta, width=60).grid(row=0, column=1, **pad)
        tk.Button(self, text="Examinar...", command=self.select_folder).grid(row=0, column=2, **pad)

        # Nombre base
        tk.Label(self, text="Nombre base:").grid(row=1, column=0, sticky="w", **pad)
        self.var_base = tk.StringVar(value=cfg["nombre_base"])
        tk.Entry(self, textvariable=self.var_base, width=40).grid(row=1, column=1, sticky="w", **pad)

        # Rango
        tk.Label(self, text="Rango a borrar (A1):").grid(row=2, column=0, sticky="w", **pad)
        self.var_rango = tk.StringVar(value=cfg["rango_borrar"])
        tk.Entry(self, textvariable=self.var_rango, width=20).grid(row=2, column=1, sticky="w", **pad)

        # Hoja
        tk.Label(self, text="Nombre de hoja (opcional):").grid(row=3, column=0, sticky="w", **pad)
        self.var_hoja = tk.StringVar(value=cfg["nombre_hoja"])
        tk.Entry(self, textvariable=self.var_hoja, width=30).grid(row=3, column=1, sticky="w", **pad)

        # Botones
        tk.Button(self, text="Guardar configuración", command=self.on_save).grid(row=4, column=0, **pad)
        tk.Button(self, text="Crear hoja de horas", command=self.on_create,
                  bg="#2e7d32", fg="white").grid(row=4, column=1, sticky="w", **pad)

        # Centrar ventana en posición 2/4
        self.center_window()

    # --------- Centrar ventana ---------
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        x = (screen_width // 2) - (width // 2)
        y = int(screen_height * 0.25)  # posición 2/4 (25%)

        self.geometry(f"{width}x{height}+{x}+{y}")

    # --------- Select folder ---------
    def select_folder(self):
        folder = filedialog.askdirectory(title="Selecciona la carpeta donde están los Excel")
        if folder:
            self.var_ruta.set(folder)

    # --------- Save config ---------
    def on_save(self):
        self.cfg["ruta_trabajo"] = self.var_ruta.get().strip()
        self.cfg["nombre_base"] = self.var_base.get().strip() or DEFAULT_BASE_NAME
        self.cfg["rango_borrar"] = self.var_rango.get().strip() or DEFAULT_RANGE
        self.cfg["nombre_hoja"] = self.var_hoja.get().strip()
        save_config(self.cfg)
        messagebox.showinfo(APP_TITLE, "Configuración guardada.")

    # --------- Create file ---------
    def on_create(self):
        try:
            self.cfg["ruta_trabajo"] = self.var_ruta.get().strip()
            self.cfg["nombre_base"] = self.var_base.get().strip() or DEFAULT_BASE_NAME
            self.cfg["rango_borrar"] = self.var_rango.get().strip() or DEFAULT_RANGE
            self.cfg["nombre_hoja"] = self.var_hoja.get().strip()

            created = create_next_file(self.cfg)
            messagebox.showinfo(APP_TITLE, f"Hoja de horas creada:\n{created}")

            # ------------------------------------------------------
            # 🔴 CERRAR AUTOMÁTICAMENTE LA APLICACIÓN DESPUÉS DE CREAR
            # Si NO quieres cierre automático → comenta la siguiente línea:
            self.destroy()
            # ------------------------------------------------------

        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Ocurrió un error:\n{e}")

def main():
    cfg = load_config()
    app = App(cfg)
    app.mainloop()

if __name__ == "__main__":
    main()